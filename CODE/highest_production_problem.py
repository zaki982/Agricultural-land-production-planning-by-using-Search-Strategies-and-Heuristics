from typing import Dict, List, Optional, Tuple
import random,math,heapq
from collections import deque
from main import Wilaya,product,Year
import networkx as nx
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from queue import Queue, LifoQueue , PriorityQueue 
from decimal import Decimal, getcontext
import csv 
import os 
import time
import tracemalloc
from filelock import FileLock
#dictionary that represent product
product_names = [
    "aubergine",
    "corn",
    "date",
    "greenpepper",
    "potatoe",
    "tomatoe",
    "wheat"
]
strategic_products=[
    "date",
    "potatoe",
    "tomatoe",
    "wheat"
]
seasons = ["winter","spring","summer","autumn"]
years = [2016,2017,2018,2019]
wilayas = [None]
wb = load_workbook("C:\python\projet\data_projet\wilaya_size.xlsx")
ws = wb.active
    
#Extract values from the specified column
size_values = [ws.cell(row=row, column=2).value for row in range(1, ws.max_row + 1)]
names = [ws.cell(row=row, column=1).value for row in range(1, ws.max_row + 1)]
#creating 48 wilaya of algeria and assign some attributes
for i in range(1,49):
    new_wilaya = Wilaya(i)
    new_wilaya.wilaya_size = size_values[i-1]
    new_wilaya.name = names[i-1]
    wilayas.append(new_wilaya)
#dictionary that contains each wilaya and it's neighbors
wilaya_neighbors = ({
wilayas[1] : [wilayas[37],wilayas[8],wilayas[32],wilayas[47],wilayas[11]],
wilayas[2] : [wilayas[27],wilayas[48],wilayas[38],wilayas[42],wilayas[44]],
wilayas[3] : [wilayas[14],wilayas[17],wilayas[47],wilayas[32],wilayas[20]],
wilayas[4] : [wilayas[43],wilayas[25],wilayas[24],wilayas[41],wilayas[12],wilayas[40],wilayas[5]],
wilayas[5] : [wilayas[19],wilayas[43],wilayas[4],wilayas[40],wilayas[7],wilayas[28],wilayas[34]],
wilayas[6] : [wilayas[15],wilayas[10],wilayas[34],wilayas[19],wilayas[18]],
wilayas[7] : [wilayas[28],wilayas[5],wilayas[40],wilayas[39],wilayas[17],wilayas[30]],
wilayas[8] : [wilayas[37],wilayas[1],wilayas[32],wilayas[45]],
wilayas[9] : [wilayas[42],wilayas[44],wilayas[26],wilayas[10],wilayas[35],wilayas[16]],
wilayas[10] : [wilayas[9],wilayas[35],wilayas[15],wilayas[6],wilayas[34],wilayas[28],wilayas[26]],
wilayas[11] : [wilayas[1],wilayas[47],wilayas[30],wilayas[33]],
wilayas[12] : [wilayas[41],wilayas[4],wilayas[40],wilayas[39]],
wilayas[13] : [wilayas[46],wilayas[22],wilayas[45]],
wilayas[14] : [wilayas[29],wilayas[48],wilayas[38],wilayas[17],wilayas[3],wilayas[32],wilayas[20],wilayas[26],wilayas[44]],
wilayas[15] : [wilayas[35],wilayas[10],wilayas[6]],
wilayas[16] : [wilayas[42],wilayas[9],wilayas[35]],
wilayas[17] : [wilayas[26],wilayas[28],wilayas[7],wilayas[39],wilayas[30],wilayas[47],wilayas[3],wilayas[14]],
wilayas[18] : [wilayas[6],wilayas[19],wilayas[43],wilayas[25],wilayas[21]],
wilayas[19] : [wilayas[6],wilayas[18],wilayas[5],wilayas[43],wilayas[34],wilayas[28]],
wilayas[20] : [wilayas[22],wilayas[29],wilayas[14],wilayas[45],wilayas[32]],
wilayas[21] : [wilayas[18],wilayas[25],wilayas[24],wilayas[23]],
wilayas[22] : [wilayas[13],wilayas[46],wilayas[31],wilayas[29],wilayas[45],wilayas[20]],
wilayas[23] : [wilayas[21],wilayas[24],wilayas[36]],
wilayas[24] : [wilayas[21],wilayas[23],wilayas[36],wilayas[41],wilayas[4],wilayas[25]],
wilayas[25] : [wilayas[18],wilayas[21],wilayas[24],wilayas[4],wilayas[43]],
wilayas[26] : [wilayas[44],wilayas[9],wilayas[10],wilayas[28],wilayas[17],wilayas[14],wilayas[38]],
wilayas[27] : [wilayas[29],wilayas[31],wilayas[48],wilayas[2]],
wilayas[28] : [wilayas[26],wilayas[10],wilayas[34],wilayas[19],wilayas[5],wilayas[7],wilayas[17]],
wilayas[29] : [wilayas[31],wilayas[22],wilayas[20],wilayas[14],wilayas[48],wilayas[27]],
wilayas[30] : [wilayas[17],wilayas[47],wilayas[39],wilayas[11],wilayas[33]],
wilayas[31] : [wilayas[22],wilayas[46],wilayas[29],wilayas[27]],
wilayas[32] : [wilayas[45],wilayas[20],wilayas[14],wilayas[3],wilayas[47],wilayas[1],wilayas[8]],
wilayas[33] : [wilayas[11],wilayas[30]],
wilayas[34] : [wilayas[10],wilayas[15],wilayas[6],wilayas[19],wilayas[28]],
wilayas[35] : [wilayas[16],wilayas[9],wilayas[10],wilayas[15]],
wilayas[36] : [wilayas[23],wilayas[24],wilayas[41]],
wilayas[37] : [wilayas[1],wilayas[8]],
wilayas[38] : [wilayas[2],wilayas[48],wilayas[44],wilayas[26],wilayas[14]],
wilayas[39] : [wilayas[40],wilayas[12],wilayas[7],wilayas[17],wilayas[30]],
wilayas[40] : [wilayas[5],wilayas[4],wilayas[12],wilayas[39],wilayas[7]],
wilayas[41] : [wilayas[24],wilayas[36],wilayas[4],wilayas[12]],
wilayas[42] : [wilayas[2],wilayas[44],wilayas[9],wilayas[16]],

wilayas[43] : [wilayas[19],wilayas[18],wilayas[25],wilayas[4],wilayas[5]],
wilayas[44] : [wilayas[2],wilayas[42],wilayas[38],wilayas[9],wilayas[26]],
wilayas[45] : [wilayas[13],wilayas[22],wilayas[20],wilayas[32],wilayas[8]],
wilayas[46] : [wilayas[31],wilayas[13],wilayas[22]],
wilayas[47] : [wilayas[32],wilayas[3],wilayas[17],wilayas[30],wilayas[11],wilayas[1]],
wilayas[48] : [wilayas[27],wilayas[2],wilayas[38],wilayas[14],wilayas[29]],
})

#dictionary that represent distance in KM between each wilaya and it's neighbors
cost = {
    'ADRAR': {
        'Neighbors': [('TINDOUF', 498), ('BECHAR', 546), ('EL BAYADH', 715), ('TIMIMOUN', 447), ('GHARDAIA', 813)]
    },
    'CHLEF': {
        'Neighbors': [('RELIZANE', 51), ('AIN DEFLA', 73), ('TIPAZA', 88), ('TISSEMSILT', 68), ('MOSTAGANEM', 83)]
    },
    'LAGHOUAT': {
        'Neighbors': [('EL BAYADH', 192), ('GHARDAIA', 168), ('TIARET', 159), ('SAIDA', 246), ('DJELFA', 84)]
    },
    'OUM EL BOUAGHI': {
        'Neighbors': [('KHENCHELA', 100), ('BATNA', 127), ('CONSTANTINE', 71), ('SOUK AHRAS', 81), ('TEBESSA', 109), ('GUELMA', 69), ('MILA', 91)]
    },
    'BATNA': {
        'Neighbors': [('OUM EL BOUAGHI', 127), ('KHENCHELA', 116), ('BORDJ BOU ARRERIDJ', 128), ('BISKRA', 61), ('MSILA', 141), ('SETIF', 99), ('MILA', 108)]        
    },
    'BEJAIA': {
        'Neighbors': [('JIJEL', 80), ('SETIF', 69), ('BOUIRA', 118), ('TIZI OUZOU', 74), ('BORDJ BOU ARRERIDJ', 81)]
    },
    'BISKRA': {
        'Neighbors': [('OUARGLA', 429), ('MSILA', 152), ('EL OUED', 213), ('KHENCHELA', 101), ('BATNA', 61), ('DJELFA', 243)]
    },
    'BECHAR': {
        'Neighbors': [('ADRAR', 546), ('TINDOUF', 591), ('NAAMA', 233), ('EL BAYADH', 324)]
    },
    'BLIDA': {
        'Neighbors': [('TIPAZA', 59), ('MEDEA', 57), ('BOUMERDES', 76), ('ALGIERS', 39), ('AIN DEFLA', 75), ('BOUIRA', 100)]
    },
    'BOUIRA': {
        'Neighbors': [('TIZI OUZOU', 57), ('BLIDA', 100), ('BORDJ BOU ARRERIDJ', 69), ('MSILA', 125), ('MEDEA', 85), ('BEJAIA', 118), ('BOUMERDES', 61)]
    },
    'TAMANRASSET': {
        'Neighbors': [('ADRAR', 617), ('ILLIZI', 520), ('GHARDAIA', 896), ('OUARGLA', 772)]
    },
    'TEBESSA': {
        'Neighbors': [('OUM EL BOUAGHI', 109), ('KHENCHELA', 123), ('EL OUED', 258), ('SOUK AHRAS', 85)]
    },
    'TLEMCEN': {
        'Neighbors': [('SIDI BEL ABBES', 83), ('AIN TEMOUCHENT', 49), ('NAAMA', 188)]
    },
    'TIARET': {
        'Neighbors': [('SAIDA', 124), ('MASCARA', 135), ('RELIZANE', 121), ('TISSEMSILT', 101), ('AIN DEFLA', 147), ('MEDEA', 175), ('DJELFA', 160), ('LAGHOUAT', 159), ('EL BAYADH', 261)]
    },
    'TIZI OUZOU': {
        'Neighbors': [('BOUMERDES', 55), ('BOUIRA', 57), ('BEJAIA', 74)]
    },
    'ALGIERS': {
        'Neighbors': [('BLIDA', 39), ('TIPAZA', 84), ('BOUMERDES', 49)]
    },
    'DJELFA': {
        'Neighbors': [('MSILA', 125), ('BISKRA', 243), ('LAGHOUAT', 84), ('OUARGLA', 498), ('GHARDAIA', 213), ('TIARET', 160), ('MEDEA', 182)]
    },
    'JIJEL': {
        'Neighbors': [('BEJAIA', 80), ('SETIF', 78), ('MILA', 3002), ('SKIKDA', 82)]
    },
    'SETIF': {
        'Neighbors': [('JIJEL', 78), ('BEJAIA', 69), ('BATNA', 99), ('OUM EL BOUAGHI', 151), ('BORDJ BOU ARRERIDJ', 67), ('MILA', 3077)]
    },
    'SAIDA': {
        'Neighbors': [('SIDI BEL ABBES', 62), ('TIARET', 124), ('MASCARA', 72), ('TLEMCEN', 143), ('EL BAYADH', 254), ('NAAMA', 193)]
    },
    'SKIKDA': {
        'Neighbors': [('JIJEL', 82), ('MILA', 2932), ('ANNABA', 79), ('GUELMA', 64), ('CONSTANTINE', 49)]
    },
    'SIDI BEL ABBES': {
        'Neighbors': [('AIN TEMOUCHENT', 94), ('ORAN', 115), ('MASCARA', 100), ('NAAMA', 164), ('SAIDA', 62), ('TLEMCEN', 83)]
    },
    'ANNABA': {
        'Neighbors': [('GUELMA', 68), ('SKIKDA', 79), ('EL TARF', 37)]
    },
    'GUELMA': {
        'Neighbors': [('ANNABA', 68), ('CONSTANTINE', 71), ('OUM EL BOUAGHI', 69), ('EL TARF', 69), ('SKIKDA', 64), ('SOUK AHRAS', 44)]
    },
    'CONSTANTINE': {
        'Neighbors': [('OUM EL BOUAGHI', 71), ('SKIKDA', 49), ('MILA', 2977), ('JIJEL', 70), ('SETIF', 110), ('GUELMA', 71)]
    },
    'MEDEA': {
        'Neighbors': [('BLIDA', 57), ('BOUIRA', 85), ('MSILA', 142), ('DJELFA', 182), ('AIN DEFLA', 86), ('TIARET', 175), ('TISSEMSILT', 108)]
    },
    'MOSTAGANEM': {
        'Neighbors': [('MASCARA', 68), ('RELIZANE', 52), ('CHLEF', 83)]
    },
    'MSILA': {
        'Neighbors': [('DJELFA', 125), ('BATNA', 141), ('BOUIRA', 125), ('BORDJ BOU ARRERIDJ', 115), ('BISKRA', 152), ('SETIF', 160), ('MEDEA', 142)]
    },
    'MASCARA': {
        'Neighbors': [('MOSTAGANEM', 68), ('RELIZANE', 77), ('TIARET', 135), ('AIN TEMOUCHENT', 126), ('SIDI BEL ABBES', 100), ('SAIDA', 72), ('ORAN', 87)]
    },
    'OUARGLA': {
        'Neighbors': [('ILLIZI', 363), ('EL OUED', 248), ('TAMANRASSET', 772), ('GHARDAIA', 343), ('DJELFA', 498), ('BISKRA', 429)]
    },
    'ORAN': {
        'Neighbors': [('AIN TEMOUCHENT', 63), ('SAIDA', 134), ('MASCARA', 87), ('SIDI BEL ABBES', 115)]
    },
    'EL BAYADH': {
        'Neighbors': [('LAGHOUAT', 192), ('GHARDAIA', 229), ('SAIDA', 254), ('ADRAR', 715), ('TIARET', 261), ('NAAMA', 195), ('BECHAR', 324)]
    },
    'ILLIZI': {
        'Neighbors': [('TAMANRASSET', 520), ('OUARGLA', 363)]
    },
    'BORDJ BOU ARRERIDJ': {
        'Neighbors': [('BATNA', 128), ('MSILA', 115), ('BOUIRA', 69), ('SETIF', 67), ('BEJAIA', 81)]
    },
    'BOUMERDES': {
        'Neighbors': [('ALGIERS', 49), ('BLIDA', 76), ('TIZI OUZOU', 55), ('BOUIRA', 61)]
    },
    'EL TARF': {
        'Neighbors': [('ANNABA', 37), ('GUELMA', 69), ('SOUK AHRAS', 63)]
    },
    'TINDOUF': {
        'Neighbors': [('ADRAR', 498), ('BECHAR', 591)]
    },
    'TISSEMSILT': {
        'Neighbors': [('RELIZANE', 83), ('AIN DEFLA', 47), ('TIARET', 101), ('CHLEF', 68), ('MEDEA', 108)]
    },
    'EL OUED': {
        'Neighbors': [('OUARGLA', 248), ('BISKRA', 213), ('KHENCHELA', 189), ('DJELFA', 385), ('TEBESSA', 258)]
    },
    'KHENCHELA': {
        'Neighbors': [('EL OUED', 189), ('BATNA', 116), ('OUM EL BOUAGHI', 100), ('TEBESSA', 123), ('SOUK AHRAS', 159), ('BISKRA', 101)]
    },
    'SOUK AHRAS': {
        'Neighbors': [('GUELMA', 44), ('OUM EL BOUAGHI', 81), ('EL TARF', 63), ('TEBESSA', 85)]
    },
    'TIPAZA': {
        'Neighbors': [('BLIDA', 59), ('ALGIERS', 84), ('CHLEF', 88), ('AIN DEFLA', 41)]
    },
    'MILA': {
        'Neighbors': [('JIJEL', 56), ('SKIKDA', 85), ('CONSTANTINE', 41), ('OUM EL BOUAGHI', 91), ('BATNA', 108), ('SETIF', 68)]
    },
    'AIN DEFLA': {
        'Neighbors': [('BLIDA', 75), ('TIPAZA', 41), ('MEDEA', 86), ('CHLEF', 73), ('TISSEMSILT', 47)]
    },
    'NAAMA': {
        'Neighbors': [('TLEMCEN', 188), ('SIDI BEL ABBES', 164), ('SAIDA', 193), ('EL BAYADH', 195), ('BECHAR', 233), ('AIN TEMOUCHENT', 231)] 
    },
    'AIN TEMOUCHENT': {
        'Neighbors': [('MASCARA', 126), ('TLEMCEN', 49), ('ORAN', 63), ('MOSTAGANEM', 157), ('SIDI BEL ABBES', 94)]
    },
    'GHARDAIA': {
        'Neighbors': [('EL BAYADH', 229), ('LAGHOUAT', 168), ('ADRAR', 813), ('OUARGLA', 343), ('DJELFA', 213), ('TAMANRASSET', 896)]
    },
    'RELIZANE': {
        'Neighbors': [('MOSTAGANEM', 52), ('CHLEF', 51), ('TIARET', 121), ('MASCARA', 77), ('TISSEMSILT', 83)]
    },
}

#dictionary which gives each region and it's corresponding wilayas
regions = {
    'north': [wilayas[16], wilayas[31], wilayas[25], wilayas[23], wilayas[13], wilayas[27], wilayas[2], wilayas[15], wilayas[6], wilayas[18], wilayas[21], wilayas[35], wilayas[42], wilayas[9], wilayas[26], wilayas[44], wilayas[48]],
    'midle': [wilayas[14], wilayas[38], wilayas[20], wilayas[29], wilayas[22], wilayas[3], wilayas[17], wilayas[28], wilayas[5], wilayas[19], wilayas[34], wilayas[4], wilayas[43], wilayas[24], wilayas[41], wilayas[12], wilayas[40], wilayas[36], wilayas[46],wilayas[10]],
    'south': [wilayas[1], wilayas[11], wilayas[37], wilayas[8], wilayas[39], wilayas[30], wilayas[47], wilayas[33], wilayas[45], wilayas[32], wilayas[7]]
}

#dictionary which contains different scales of suitability for each product according to the region
agricultural_suitability = {
    'wheat': {'north': 0.9, 'midle': 0.5, 'south': 0.1},
    'corn': {'north': 0.9, 'midle': 0.5, 'south': 0.1},
    'date': {'north': 0.1, 'midle': 0.5, 'south': 0.9},
    'potatoe': {'north': 0.9, 'midle': 0.5, 'south': 0.1},
    'tomatoe': {'north': 0.9, 'midle': 0.5, 'south': 0.5},
    'greenpepper': {'north': 0.9, 'midle': 0.5, 'south': 0.5},
    'aubergine': {'north': 0.9, 'midle': 0.5, 'south': 0.5}
}

quotient = {
    'aubergine': {
        2016: 2.2637 * 10**-7,
        2017: 1.8152 * 10**-7,
        2018: 1.1593 * 10**-7,
        2019: 1.0842 * 10**-7
    },
    'corn': {
        2016: 4.4911 * 10**-8,
        2017: 1.3455 * 10**-8,
        2018: 2.6864 * 10**-8,
        2019: 1.0530 * 10**-8
    },
    'date': {
        2016: 1.2282 * 10**-9,
        2017: 1.2222 * 10**-9,
        2018: 1.2138 * 10**-9,
        2019: 1.2073 * 10**-9
    },
    'greenpepper': {
        2016: 1.1318 * 10**-7,
        2017: 1.1972 * 10**-7,
        2018: 9.1636 * 10**-8,
        2019: 8.4787 * 10**-8
    },
    'potatoe': {
        2016: 4.2614 * 10**-9,
        2017: 5.2354 * 10**-9,
        2018: 5.0619 * 10**-9,
        2019: 4.9524 * 10**-9
    },
    'tomatoe': {
        2016: 4.3536 * 10**-8,
        2017: 4.5017 * 10**-8,
        2018: 3.8772 * 10**-8,
        2019: 3.8075 * 10**-8
    },
    'wheat': {
        2016: 2.5878 * 10**-9,
        2017: 7.1662 * 10**-10,
        2018: 7.1662 * 10**-10,
        2019: 2.4129 * 10**-9
    }
}


highest_production_wilaya = {
    'aubergine': {
        2016: 'BISKRA',
        2017: 'BISKRA',
        2018: 'BISKRA',
        2019: 'BISKRA'
    },
    'corn': {
        2016: 'ADRAR',
        2017: 'ADRAR',
        2018: 'ADRAR',
        2019: 'ADRAR'
    },
    'date': {
        2016: 'BISKRA',
        2017: 'BISKRA',
        2018: 'BISKRA',
        2019: 'BISKRA'
    },
    'greenpepper': {
        2016: 'BISKRA',
        2017: 'BISKRA',
        2018: 'BISKRA',
        2019: 'BISKRA'
    },
    'potatoe': {
        2016: 'EL-OUED',
        2017: 'EL-OUED',
        2018: 'EL-OUED',
        2019: 'EL-OUED'
    },
    'tomatoe': {
        2016: 'BISKRA',
        2017: 'BISKRA',
        2018: 'BISKRA',
        2019: 'BISKRA'
    },
    'wheat': {
        2016: 'SETIF',
        2017: 'TIARET',
        2018: 'TIARET',
        2019: 'SETIF'
    }
}


highest_values = {}

#find highest quotient (used in heuristic)
for prod, ye in quotient.items():
    highest_value = max(ye.values())
    highest_values[prod] = highest_value

#implementation of class Node
class Node:
    def __init__(self, wilaya: Wilaya, parent=None, action=None, path_cost=0, heuristic = 0):
        self.wilaya = wilaya
        self.parent = parent
        self.action = action
        self.path_cost = path_cost
        self.heuristic = heuristic

    def __hash__(self):
        return hash(self.wilaya)
    #checks if node eqal to other node
    def __eq__(self, other):
        return isinstance(other, Node) and self.wilaya == other.wilaya
    #chekcs if node is less than another node
    def __lt__(self, other):
        return (self.heuristic + self.path_cost< other.heuristic + other.path_cost)
    #return name of the node ( wilaya ) as string
    def __str__(self):
        return self.wilaya.get_wilaya_name() 

#implementation of production problem class
class production_problem:
    def __init__(self, initial_wilaya: Wilaya, product: product,year):
        self.initial_wilaya = initial_wilaya
        self.product = product
        self.highest_production = -1  
        self.wilayet = wilayas
        self.region=regions
        self.nei = wilaya_neighbors
        # Create a graph that has wilaya as nodes and link between them according to wilayas_neighbors dictionary
        self.G = nx.Graph(self.nei)
        #draw graph :
        #nx.draw(self.G, with_labels=True)
        #plt.show()
        self.year = year

    
    def get_initial_state(self):
        return self.initial_wilaya

    #function that return neighbors of a current wilaya
    def possible_directions(self, wilaya):
        if wilaya in self.G:
            return list(self.G.neighbors(wilaya))
        else:
            return []

    #boolean function that checks whether a node(wilaya) is the goal or not in case of bfs or dfs search
    def goal_test_bfs_dfs(self, wilaya):
        product_name = self.product.get_name()
        if product_name is None:
            print("Product name is not set correctly.")
            return False
        year = self.year
        highest_prod_wilaya = highest_production_wilaya.get(product_name, {}).get(year)
        if highest_prod_wilaya is None:
            print("Highest production wilaya not found for the given product and year.")
            return False
        if wilaya.get_wilaya_name().upper() == highest_prod_wilaya:
            return True
        else:
            return False
    
    #function return the distance between a wilaya and a one of it's neighbors
    def get_step_cost(self, start_wilaya , end_wilaya):
      z = 0
      for i in range(0, len(cost[start_wilaya]['Neighbors'])):
            if cost[start_wilaya]['Neighbors'][i][0] == end_wilaya:
               z = i 
               break
      return cost[start_wilaya]['Neighbors'][z][1]
    
    #heuristic function for highest_production problem
    def heuristic_abzk(self,Wilaya,product,year):
        z = 0
        if Wilaya in regions["north"]:
           z = "north"
        elif Wilaya in regions["south"]:
           z = "south"
        elif Wilaya in regions["midle"]:
            z = "midle"
        if z == 0:
            print("wilaya not found")
            z = "midle"
        produit = Wilaya.wilaya_size*self.wilayet[Wilaya.get_wilaya_number()].wilaya_products[self.product.get_name()].yearly_info.yearly_production_data[self.year][0]*agricultural_suitability[self.product.get_name()][z]
        quotient = 1/produit
        return quotient
    
    #function which represent the step_cost g(n)
    def g(self,wilaya,product):
        z = 0
        if wilaya in regions["north"]:
           z = "north"
        elif wilaya in regions["south"]:
           z = "south"
        elif wilaya in regions["midle"]:
            z = "midle"
        if z == 0:
            print("wilaya not found")
            z = "midle" #if wilaya not found assign z to midle by default
        return 1 - agricultural_suitability[product][z]
    
    #function which test whether a node is the goal or not in case of A* search
    def goal_test(self,node):
        return self.heuristic_abzk(node.wilaya,self.product.get_name(),self.year) <= highest_values[self.product.get_name()] 
    
    #function which return the best neighbor of any wilaya according to their production level
    def best_neighbors(self,node):
        successors = self.possible_directions(node.wilaya)
        best_neighbor = None
        highest_prod = -1
        for successor in successors:
            prod =  successor.wilaya_products[self.product.get_name()].yearly_info.yearly_production_data[self.year][1]
            if prod >= highest_prod:
                highest_prod = prod
                best_neighbor = successor
        b = Node(best_neighbor,node,highest_prod)
        return b

#breath first search algorithm
def bfs(problem):
    start = problem.get_initial_state()
    visited = set()
    node = Node(problem.get_initial_state())
    queue = deque([node])

    while queue:
        node = queue.popleft()
        while node in visited:
            if not len(queue)==0:
                node = queue.popleft()
            else:
                return None
          #print path
        if problem.goal_test_bfs_dfs(node.wilaya):
           return node,node.wilaya.wilaya_products[problem.product.get_name()].yearly_info.yearly_production_data[problem.year][1]
        visited.add(node)
        #print("Visiting node:", node.wilaya.get_wilaya_number())          #print path
        for child in problem.possible_directions(node.wilaya):
            child = Node(child, node)
            queue.append(child)
            node =queue.popleft()
            
            if not node in visited:
                queue.append(node)
        
    return None
# depth first search algorithm
def dfs(problem):
    start = problem.get_initial_state()
    visited = set()
    node = Node(problem.get_initial_state())
    frontier = [node]


    while frontier:
     node = frontier.pop()
     #print("Visiting node:", node.wilaya.get_wilaya_number())  # to print the path
     if problem.goal_test_bfs_dfs(node.wilaya):
        return node.wilaya.get_wilaya_name(),node.wilaya.wilaya_products[problem.product.get_name()].yearly_info.yearly_production_data[problem.year][1]
     visited.add(node)
     for child in problem.possible_directions(node.wilaya):
        child = Node(child, node)
        frontier.append(child)
        node =frontier.pop()
        if node not in visited:
            frontier.append(node)
    return None

# hill climbing / type : steepest acent 
def hill_climbing(problem):
    currentState = Node(problem.get_initial_state())
    while True:
        neighbor_state = problem.best_neighbors(currentState)
        if neighbor_state.wilaya.wilaya_products[problem.product.get_name()].yearly_info.yearly_production_data[problem.year][1] < currentState.wilaya.wilaya_products[problem.product.get_name()].yearly_info.yearly_production_data[problem.year][1]:
            return currentState.wilaya.get_wilaya_name()
        else :
            currentState = neighbor_state

#A* search algorithm 
def AStarSearch(problem):
    node = Node(problem.get_initial_state())
    closed = set()
    opened = PriorityQueue()
    opened.put((0, node))  # Priority queue with (cost, node)

    while not opened.empty():
        current_cost, current_node = opened.get()
        
        if problem.goal_test(current_node):
            print('goal:', current_node.wilaya.get_wilaya_name())
            return current_node
        
        closed.add(current_node)

        # Expand the current node neighbors
        for neighbor in problem.nei[current_node.wilaya]:
            if neighbor in closed:
                continue
            
            g_cost = problem.g(neighbor, problem.product.get_name())
            h_cost = problem.heuristic_abzk(neighbor, problem.product.get_name(), problem.year)
            total_cost = current_cost + g_cost + h_cost

            neighbor_node = Node(neighbor, heuristic=h_cost, path_cost=g_cost, parent=current_node)

            # Check if the neighbor is already in the priority queue with a higher cost
            in_opened = False
            for item in opened.queue:
                if item[1] == neighbor_node and item[0] <= total_cost:
                    in_opened = True
                    break
            
            if not in_opened:
                opened.put((total_cost, neighbor_node))
    
    return None
w = wilayas[31]
p = product("aubergine")
y = 2016
p = production_problem(w,p,y)
# Directory paths
csv_directory1 = "C:/python/projet/data_projet/data_2016-2019_csv/"
csv_directory = "C:/python/projet/data_projet/prices_data_2016-2019_csv/"
csv_directory3 = "C:/python/projet/data_projet/sorted_data_csv/"
csv_directory2 = "C:\python\projet\data_projet\consumption.csv"
csv_directory4 = "C:\python\projet\data_projet\production.csv"

# Lock file paths
lock_file_path1 = lambda product_name, year: f"{csv_directory1}{product_name}_{year}.csv.lock"
lock_file_path = lambda product_name, year: f"{csv_directory}{product_name}_{year}.csv.lock"
lock_file_path3 = lambda product_name, year: f"{csv_directory3}{product_name}_{year}.csv.lock"
lock_file_path2 = f"{csv_directory2}.csv.lock"
lock_file_path4 = f"{csv_directory4}.csv.lock"

# Read and process production data
for product_name in product_names:
    for year in years:
        number=1
        file_path1 = f"{csv_directory1}{product_name}_{year}.csv"
        lock = FileLock(lock_file_path1(product_name, year))

        with lock, open(file_path1, newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile, delimiter=',')
            #print(f"Reading data for {product_name} - {year}:")
            file_processed = False
            for row in reader:
                try:
                    if len(row) < 4:
                        print(f"Warning: Incomplete data in CSV file: {row}")
                        continue
                    wilaya_name = row[0]
                    land_size = float(row[1])
                    yearly_production = float(row[2])
                    productivity = float(row[3])
                    if land_size == 0 and yearly_production == 0 and productivity ==0:
                        land_size = 1
                        yearly_production = 1
                        productivity = 1
                    else:
                        land_size = float(row[1])
                        yearly_production = float(row[2])
                        productivity = float(row[3])
                    #print(f"Wilaya: {wilaya_name} , land_size: {land_size} , yearly_production: {yearly_production} , productivity: {productivity} ")

                    p.wilayet[number].wilaya_products[product_name].yearly_info.set_yearly_production(year, land_size, yearly_production, productivity)
                    number += 1
                    file_processed = True
                except ValueError as e:
                    print(f"Error converting row to floats: {row}")
                    print(e)

            if not file_processed:
                print(f"No data found for {product_name} - {year}")
            #print("-" * 40)

# Read and process price data
for product_name in product_names:
    for year in years:
        number=1
        file_path = f"{csv_directory}{product_name}_{year}.csv"
        lock = FileLock(lock_file_path(product_name, year))

        with lock, open(file_path, newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile, delimiter=',')
            file_processed = False
            for row in reader:
                try:
                    if len(row) < 5:
                        print(f"Warning: Incomplete data in CSV file: {row}")
                        continue

                    wilaya_name = row[0]
                    season_prices = [float(price) for price in row[1:5]]

                    for i, season in enumerate(seasons):
                        p.wilayet[number].wilaya_products[product_name].yearly_info.set_prices_per_season(year, season_prices[i], season)
                    number += 1
                    file_processed = True
                except ValueError as e:
                    print(f"Error converting row to floats: {row}")
                    print(e)

            if not file_processed:
                print(f"No data found for {product_name} - {year}")
# Directory paths

csv_directory2 = "C:\python\projet\data_projet\consumption.csv"
csv_directory4 = "C:\python\projet\data_projet\production.csv"

lock_file_path2 = f"{csv_directory2}.csv.lock"
lock_file_path4 = f"{csv_directory4}.csv.lock"



# Dictionary to store product instances
products_dict = {name: product(name) for name in product_names}

# Read and process consumption data
lock = FileLock(lock_file_path2)
with lock, open(csv_directory2, newline='', encoding='utf-8') as csvfile:
    reader = csv.reader(csvfile, delimiter=',')
    for row in reader:
        if row:
            product_name = row[0]
            consumption_data = [int(x) for x in row[1:]]
            current_product = products_dict[product_name]
            for i in range(4):
                year = years[i]
                consumption = consumption_data[i]
                current_product.yearly_info.set_country_yearly_consumption(year, consumption)

# Read and process production data
lock = FileLock(lock_file_path4)
with lock, open(csv_directory4, newline='', encoding='utf-8') as csvfile:
    reader = csv.reader(csvfile, delimiter=',')
    for row in reader:
        if row:
            product_name = row[0]
            production_data = [int(x) for x in row[1:]]
            current_product = products_dict[product_name]
            for i in range(4):
                year = years[i]
                production = production_data[i]
                current_product.yearly_info.set_country_yearly_production(year, production)
'''
IF YOU WANT TO RUN THE CODE,THE PRINTED STATEMENTS ARE FOR SOME SEARCH ALGOS AND IT WILL BE REAPETED 100 TIMES
BECAUSE OF THE MEASURING PROGRAMS BELOW HENCE IT DOES NOT REPRESENT THE PATHS VISITED FOR ONE ATTEPMTED SEARCH.
THANK YOU,
'''
# Measure time in microseconds with multiple trials
def measure_time_bfs(trials=100):
    times = []
    for _ in range(trials):
        start_time = time.perf_counter()
        bfs(p)
        end_time = time.perf_counter()
        time_taken = (end_time - start_time) * 1e6  # convert to microseconds
        times.append(time_taken)
    avg_time = sum(times) / len(times)
    return avg_time
# Measure space in bytes
tracemalloc.start()
bfs(p)
current, peak = tracemalloc.get_traced_memory()
tracemalloc.stop()

avg_time = measure_time_bfs()
print(f"Average time taken over {100} trials: {avg_time:.2f} µs")
print(f"Current memory usage: {current} bytes; Peak: {peak} bytes")


# Measure time in microseconds with multiple trials
def measure_time_dfs(trials=100):
    times = []
    for _ in range(trials):
        start_time = time.perf_counter()
        dfs(p)
        end_time = time.perf_counter()
        time_taken = (end_time - start_time) * 1e6  # convert to microseconds
        times.append(time_taken)
    avg_time = sum(times) / len(times)
    return avg_time

# Measure space in bytes
tracemalloc.start()
dfs(p)
current, peak = tracemalloc.get_traced_memory()
tracemalloc.stop()

avg_time = measure_time_dfs()
print(f"Average time taken over {100} trials: {avg_time:.2f} µs")
print(f"Current memory usage: {current} bytes; Peak: {peak} bytes")

# Measure time in microseconds with multiple trials
def measure_time_hill_climbing(trials=100):
    times = []
    for _ in range(trials):
        start_time = time.perf_counter()
        hill_climbing(p)
        end_time = time.perf_counter()
        time_taken = (end_time - start_time) * 1e6  # convert to microseconds
        times.append(time_taken)
    avg_time = sum(times) / len(times)
    return avg_time

# Measure space in bytes
tracemalloc.start()
hill_climbing(p)
current, peak = tracemalloc.get_traced_memory()
tracemalloc.stop()

avg_time = measure_time_hill_climbing()
print(f"Average time taken over {100} trials: {avg_time:.2f} µs")
print(f"Current memory usage: {current} bytes; Peak: {peak} bytes")
