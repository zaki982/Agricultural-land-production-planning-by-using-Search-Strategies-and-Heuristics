from typing import Dict, List, Optional, Tuple
import random,math,heapq
from collections import deque
from main import Wilaya,product,Year
import networkx as nx
import pandas as pd
import matplotlib.pyplot as plt
import zipfile
from openpyxl import load_workbook
import csv
import os
import math
from queue import Queue, LifoQueue , PriorityQueue 
from decimal import Decimal, getcontext
from filelock import FileLock
import time
import tracemalloc

#dictionary that represent product
product_names = [
    "aubergine",
    "corn",
    "date",
    "greenpepper",
    "potatoe",
    "tomatoe",
    "wheat"
]
seasons = ["winter","spring","summer","autumn"]
years = [2016,2017,2018,2019]



wilayas = [None]
wb = load_workbook("C:\python\projet\data_projet\wilaya_size.xlsx")
ws = wb.active
    # Extract values from the specified column
size_values = [ws.cell(row=row, column=2).value for row in range(1, ws.max_row + 1)]
names = [ws.cell(row=row, column=1).value for row in range(1, ws.max_row + 1)]
#creating 48 wilaya of algeria and assign some attributes
for i in range(1,49):
    new_wilaya = Wilaya(i)
    new_wilaya.wilaya_size = size_values[i-1]
    new_wilaya.name = names[i-1]
    wilayas.append(new_wilaya)
#dictionary that contains each wilaya and it's neighbors
wilaya_neighbors = ({
wilayas[1] : [wilayas[37],wilayas[8],wilayas[32],wilayas[47],wilayas[11]],
wilayas[2] : [wilayas[27],wilayas[48],wilayas[38],wilayas[42],wilayas[44]],
wilayas[3] : [wilayas[14],wilayas[17],wilayas[47],wilayas[32]],
wilayas[4] : [wilayas[43],wilayas[25],wilayas[24],wilayas[41],wilayas[12],wilayas[40],wilayas[5]],
wilayas[5] : [wilayas[19],wilayas[43],wilayas[4],wilayas[40],wilayas[7],wilayas[28]],
wilayas[6] : [wilayas[15],wilayas[10],wilayas[34],wilayas[19],wilayas[18]],
wilayas[7] : [wilayas[28],wilayas[5],wilayas[40],wilayas[39],wilayas[17]],
wilayas[8] : [wilayas[37],wilayas[1],wilayas[32],wilayas[45]],
wilayas[9] : [wilayas[42],wilayas[44],wilayas[26],wilayas[10],wilayas[35],wilayas[16]],
wilayas[10] : [wilayas[9],wilayas[35],wilayas[15],wilayas[6],wilayas[34],wilayas[28],wilayas[26]],
wilayas[11] : [wilayas[1],wilayas[47],wilayas[30],wilayas[33]],
wilayas[12] : [wilayas[41],wilayas[4],wilayas[40],wilayas[39]],
wilayas[13] : [wilayas[46],wilayas[22],wilayas[45]],
wilayas[14] : [wilayas[29],wilayas[48],wilayas[38],wilayas[17],wilayas[3],wilayas[32],wilayas[20]],
wilayas[15] : [wilayas[35],wilayas[10],wilayas[6]],
wilayas[16] : [wilayas[42],wilayas[9],wilayas[35]],
wilayas[17] : [wilayas[26],wilayas[28],wilayas[7],wilayas[39],wilayas[30],wilayas[47],wilayas[3],wilayas[14]],
wilayas[18] : [wilayas[6],wilayas[19],wilayas[43],wilayas[25],wilayas[21]],
wilayas[19] : [wilayas[6],wilayas[18],wilayas[5],wilayas[43],wilayas[34],wilayas[28]],
wilayas[20] : [wilayas[22],wilayas[29],wilayas[14],wilayas[45],wilayas[32]],
wilayas[21] : [wilayas[18],wilayas[25],wilayas[24],wilayas[23]],
wilayas[22] : [wilayas[13],wilayas[46],wilayas[31],wilayas[29],wilayas[45],wilayas[20]],
wilayas[23] : [wilayas[21],wilayas[24],wilayas[36]],
wilayas[24] : [wilayas[21],wilayas[23],wilayas[36],wilayas[41],wilayas[4],wilayas[25]],
wilayas[25] : [wilayas[18],wilayas[21],wilayas[24],wilayas[4],wilayas[43]],
wilayas[26] : [wilayas[44],wilayas[9],wilayas[10],wilayas[28],wilayas[17],wilayas[14],wilayas[38]],
wilayas[27] : [wilayas[29],wilayas[31],wilayas[48],wilayas[2]],
wilayas[28] : [wilayas[26],wilayas[10],wilayas[34],wilayas[19],wilayas[5],wilayas[7],wilayas[17]],
wilayas[29] : [wilayas[31],wilayas[22],wilayas[20],wilayas[14],wilayas[48],wilayas[27]],
wilayas[30] : [wilayas[17],wilayas[47],wilayas[39],wilayas[11],wilayas[33]],
wilayas[31] : [wilayas[22],wilayas[46],wilayas[29],wilayas[27]],
wilayas[32] : [wilayas[45],wilayas[20],wilayas[14],wilayas[3],wilayas[47],wilayas[1],wilayas[8]],
wilayas[33] : [wilayas[11],wilayas[30]],
wilayas[34] : [wilayas[10],wilayas[15],wilayas[6],wilayas[19],wilayas[28]],
wilayas[35] : [wilayas[16],wilayas[9],wilayas[10],wilayas[15]],
wilayas[36] : [wilayas[23],wilayas[24],wilayas[41]],
wilayas[37] : [wilayas[1],wilayas[8]],
wilayas[38] : [wilayas[2],wilayas[48],wilayas[44],wilayas[26],wilayas[14]],
wilayas[39] : [wilayas[40],wilayas[12],wilayas[7],wilayas[17],wilayas[30]],
wilayas[40] : [wilayas[5],wilayas[4],wilayas[12],wilayas[39],wilayas[7]],
wilayas[41] : [wilayas[24],wilayas[36],wilayas[4],wilayas[12]],
wilayas[42] : [wilayas[2],wilayas[44],wilayas[9],wilayas[16]],
wilayas[43] : [wilayas[19],wilayas[18],wilayas[25],wilayas[4],wilayas[5]],
wilayas[44] : [wilayas[2],wilayas[42],wilayas[38],wilayas[9],wilayas[26]],
wilayas[45] : [wilayas[13],wilayas[22],wilayas[20],wilayas[32],wilayas[8]],
wilayas[46] : [wilayas[31],wilayas[13],wilayas[22]],
wilayas[47] : [wilayas[32],wilayas[3],wilayas[17],wilayas[30],wilayas[11],wilayas[1]],
wilayas[48] : [wilayas[27],wilayas[2],wilayas[38],wilayas[14],wilayas[29]],
})
#dictionary that represent distance in KM between each wilaya and it's neighbors
cost = {
    'ADRAR': {
        'Neighbors': [('TINDOUF', 498), ('BECHAR', 546), ('EL BAYADH', 715), ('TIMIMOUN', 447), ('GHARDAIA', 813)]
    },
    'CHLEF': {
        'Neighbors': [('RELIZANE', 51), ('AIN DEFLA', 73), ('TIPAZA', 88), ('TISSEMSILT', 68), ('MOSTAGANEM', 83)]
    },
    'LAGHOUAT': {
        'Neighbors': [('EL BAYADH', 192), ('GHARDAIA', 168), ('TIARET', 159), ('SAIDA', 246), ('DJELFA', 84)]
    },
    'OUM EL BOUAGHI': {
        'Neighbors': [('KHENCHELA', 100), ('BATNA', 127), ('CONSTANTINE', 71), ('SOUK AHRAS', 81), ('TEBESSA', 109), ('GUELMA', 69), ('MILA', 91)]
    },
    'BATNA': {
        'Neighbors': [('OUM EL BOUAGHI', 127), ('KHENCHELA', 116), ('BORDJ BOU ARRERIDJ', 128), ('BISKRA', 61), ('MSILA', 141), ('SETIF', 99), ('MILA', 108)]        
    },
    'BEJAIA': {
        'Neighbors': [('JIJEL', 80), ('SETIF', 69), ('BOUIRA', 118), ('TIZI OUZOU', 74), ('BORDJ BOU ARRERIDJ', 81)]
    },
    'BISKRA': {
        'Neighbors': [('OUARGLA', 429), ('MSILA', 152), ('EL OUED', 213), ('KHENCHELA', 101), ('BATNA', 61), ('DJELFA', 243)]
    },
    'BECHAR': {
        'Neighbors': [('ADRAR', 546), ('TINDOUF', 591), ('NAAMA', 233), ('EL BAYADH', 324)]
    },
    'BLIDA': {
        'Neighbors': [('TIPAZA', 59), ('MEDEA', 57), ('BOUMERDES', 76), ('ALGIERS', 39), ('AIN DEFLA', 75), ('BOUIRA', 100)]
    },
    'BOUIRA': {
        'Neighbors': [('TIZI OUZOU', 57), ('BLIDA', 100), ('BORDJ BOU ARRERIDJ', 69), ('MSILA', 125), ('MEDEA', 85), ('BEJAIA', 118), ('BOUMERDES', 61)]
    },
    'TAMANRASSET': {
        'Neighbors': [('ADRAR', 617), ('ILLIZI', 520), ('GHARDAIA', 896), ('OUARGLA', 772)]
    },
    'TEBESSA': {
        'Neighbors': [('OUM EL BOUAGHI', 109), ('KHENCHELA', 123), ('EL OUED', 258), ('SOUK AHRAS', 85)]
    },
    'TLEMCEN': {
        'Neighbors': [('SIDI BEL ABBES', 83), ('AIN TEMOUCHENT', 49), ('NAAMA', 188)]
    },
    'TIARET': {
        'Neighbors': [('SAIDA', 124), ('MASCARA', 135), ('RELIZANE', 121), ('TISSEMSILT', 101), ('AIN DEFLA', 147), ('MEDEA', 175), ('DJELFA', 160), ('LAGHOUAT', 159), ('EL BAYADH', 261)]
    },
    'TIZI OUZOU': {
        'Neighbors': [('BOUMERDES', 55), ('BOUIRA', 57), ('BEJAIA', 74)]
    },
    'ALGIERS': {
        'Neighbors': [('BLIDA', 39), ('TIPAZA', 84), ('BOUMERDES', 49)]
    },
    'DJELFA': {
        'Neighbors': [('MSILA', 125), ('BISKRA', 243), ('LAGHOUAT', 84), ('OUARGLA', 498), ('GHARDAIA', 213), ('TIARET', 160), ('MEDEA', 182)]
    },
    'JIJEL': {
        'Neighbors': [('BEJAIA', 80), ('SETIF', 78), ('MILA', 3002), ('SKIKDA', 82)]
    },
    'SETIF': {
        'Neighbors': [('JIJEL', 78), ('BEJAIA', 69), ('BATNA', 99), ('OUM EL BOUAGHI', 151), ('BORDJ BOU ARRERIDJ', 67), ('MILA', 3077)]
    },
    'SAIDA': {
        'Neighbors': [('SIDI BEL ABBES', 62), ('TIARET', 124), ('MASCARA', 72), ('TLEMCEN', 143), ('EL BAYADH', 254), ('NAAMA', 193)]
    },
    'SKIKDA': {
        'Neighbors': [('JIJEL', 82), ('MILA', 2932), ('ANNABA', 79), ('GUELMA', 64), ('CONSTANTINE', 49)]
    },
    'SIDI BEL ABBES': {
        'Neighbors': [('AIN TEMOUCHENT', 94), ('ORAN', 115), ('MASCARA', 100), ('NAAMA', 164), ('SAIDA', 62), ('TLEMCEN', 83)]
    },
    'ANNABA': {
        'Neighbors': [('GUELMA', 68), ('SKIKDA', 79), ('EL TARF', 37)]
    },
    'GUELMA': {
        'Neighbors': [('ANNABA', 68), ('CONSTANTINE', 71), ('OUM EL BOUAGHI', 69), ('EL TARF', 69), ('SKIKDA', 64), ('SOUK AHRAS', 44)]
    },
    'CONSTANTINE': {
        'Neighbors': [('OUM EL BOUAGHI', 71), ('SKIKDA', 49), ('MILA', 2977), ('JIJEL', 70), ('SETIF', 110), ('GUELMA', 71)]
    },
    'MEDEA': {
        'Neighbors': [('BLIDA', 57), ('BOUIRA', 85), ('MSILA', 142), ('DJELFA', 182), ('AIN DEFLA', 86), ('TIARET', 175), ('TISSEMSILT', 108)]
    },
    'MOSTAGANEM': {
        'Neighbors': [('MASCARA', 68), ('RELIZANE', 52), ('CHLEF', 83)]
    },
    'MSILA': {
        'Neighbors': [('DJELFA', 125), ('BATNA', 141), ('BOUIRA', 125), ('BORDJ BOU ARRERIDJ', 115), ('BISKRA', 152), ('SETIF', 160), ('MEDEA', 142)]
    },
    'MASCARA': {
        'Neighbors': [('MOSTAGANEM', 68), ('RELIZANE', 77), ('TIARET', 135), ('AIN TEMOUCHENT', 126), ('SIDI BEL ABBES', 100), ('SAIDA', 72), ('ORAN', 87)]
    },
    'OUARGLA': {
        'Neighbors': [('ILLIZI', 363), ('EL OUED', 248), ('TAMANRASSET', 772), ('GHARDAIA', 343), ('DJELFA', 498), ('BISKRA', 429)]
    },
    'ORAN': {
        'Neighbors': [('AIN TEMOUCHENT', 63), ('SAIDA', 134), ('MASCARA', 87), ('SIDI BEL ABBES', 115)]
    },
    'EL BAYADH': {
        'Neighbors': [('LAGHOUAT', 192), ('GHARDAIA', 229), ('SAIDA', 254), ('ADRAR', 715), ('TIARET', 261), ('NAAMA', 195), ('BECHAR', 324)]
    },
    'ILLIZI': {
        'Neighbors': [('TAMANRASSET', 520), ('OUARGLA', 363)]
    },
    'BORDJ BOU ARRERIDJ': {
        'Neighbors': [('BATNA', 128), ('MSILA', 115), ('BOUIRA', 69), ('SETIF', 67), ('BEJAIA', 81)]
    },
    'BOUMERDES': {
        'Neighbors': [('ALGIERS', 49), ('BLIDA', 76), ('TIZI OUZOU', 55), ('BOUIRA', 61)]
    },
    'EL TARF': {
        'Neighbors': [('ANNABA', 37), ('GUELMA', 69), ('SOUK AHRAS', 63)]
    },
    'TINDOUF': {
        'Neighbors': [('ADRAR', 498), ('BECHAR', 591)]
    },
    'TISSEMSlLT': {
        'Neighbors': [('RELIZANE', 83), ('AIN DEFLA', 47), ('TIARET', 101), ('CHLEF', 68), ('MEDEA', 108)]
    },
    'EL OUED': {
        'Neighbors': [('OUARGLA', 248), ('BISKRA', 213), ('KHENCHELA', 189), ('DJELFA', 385), ('TEBESSA', 258)]
    },
    'KHENCHELA': {
        'Neighbors': [('EL OUED', 189), ('BATNA', 116), ('OUM EL BOUAGHI', 100), ('TEBESSA', 123), ('SOUK AHRAS', 159), ('BISKRA', 101)]
    },
    'SOUK AHRAS': {
        'Neighbors': [('GUELMA', 44), ('OUM EL BOUAGHI', 81), ('EL TARF', 63), ('TEBESSA', 85)]
    },
    'TIPAZA': {
        'Neighbors': [('BLIDA', 59), ('ALGIERS', 84), ('CHLEF', 88), ('AIN DEFLA', 41)]
    },
    'MILA': {
        'Neighbors': [('JIJEL', 56), ('SKIKDA', 85), ('CONSTANTINE', 41), ('OUM EL BOUAGHI', 91), ('BATNA', 108), ('SETIF', 68)]
    },
    'AIN DEFLA': {
        'Neighbors': [('BLIDA', 75), ('TIPAZA', 41), ('MEDEA', 86), ('CHLEF', 73), ('TISSEMSILT', 47)]
    },
    'NAAMA': {
        'Neighbors': [('TLEMCEN', 188), ('SIDI BEL ABBES', 164), ('SAIDA', 193), ('EL BAYADH', 195), ('BECHAR', 233), ('AIN TEMOUCHENT', 231)] 
    },
    'AIN TEMOUCHENT': {
        'Neighbors': [('MASCARA', 126), ('TLEMCEN', 49), ('ORAN', 63), ('MOSTAGANEM', 157), ('SIDI BEL ABBES', 94)]
    },
    'GHARDAIA': {
        'Neighbors': [('EL BAYADH', 229), ('LAGHOUAT', 168), ('ADRAR', 813), ('OUARGLA', 343), ('DJELFA', 213), ('TAMANRASSET', 896)]
    },
    'RELIZANE': {
        'Neighbors': [('MOSTAGANEM', 52), ('CHLEF', 51), ('TIARET', 121), ('MASCARA', 77), ('TISSEMSILT', 83)]
    },
}

#values of mean production for each production in the 4 years
mean_production = {
    "aubergine": {
        2016: 2500.375,
        2017: 32474.583333333332,
        2018: 37837.125,
        2019: 38363.6875
    },
    "corn": {
        2016: 768.0208333333334,
        2017: 548.6458333333334,
        2018: 1148.4375,
        2019: 1332.5354166666666
    },
    "date": {
        2016: 214519.9375,
        2017: 220533.0625,
        2018: 228062.5,
        2019: 236671.85416666666
    },
    "greenpepper": {
        2016: 70751.97916666667,
        2017: 71589.5,
        2018: 71190.66666666667,
        2019: 76434.54166666667
    },
    "potatoe": {
        2016: 991599.3125,
        2017: 959667.1875,
        2018: 969442.1458333334,
        2019: 1045885.4166666666
    },
    "tomatoe": {
        2016: 266785.4583333333,
        2017: 228392.875,
        2018: 272863.5833333333,
        2019: 307891.4166666667
    },
    "wheat": {
        2016: 403670.2916666667,
        2017: 414782.7083333333,
        2018: 662087.6458333334,
        2019: 668493.2916666666
    }
}

quotients = {
    "aubergine": {
        2016: 0.00901555852022788,
        2017: 0.1088618662912183,
        2018: 0.09533163265306123,
        2019: 0.09868599596136285
    },
    "corn": {
        2016: 0.047596729879358785,
        2017: 0.04971869808186075,
        2018: 0.036086017282010996,
        2019: 0.03435609283418416
    },
    "date": {
        2016: 0.05006428111340081,
        2017: 0.050349542960899225,
        2018: 0.04964513456457258,
        2019: 0.050110492095419575
    },
    "greenpepper": {
        2016: 0.11608390484940963,
        2017: 0.11444432010742718,
        2018: 0.09093200493890238,
        2019: 0.08810542706250085
    },
    "potatoe": {
        2016: 0.09050742173238409,
        2017: 0.0832321931916739,
        2018: 0.08533821706279343,
        2019: 0.08615201125755079
    },
    "tomatoe": {
        2016: 0.08684676905681915,
        2017: 0.07238939196085019,
        2018: 0.07558443327193066,
        2019: 0.08402507905145531
    },
    "wheat": {
        2016: 0.2004349075690559,
        2017: 0.21784806109943977,
        2018: 0.26611239784297963,
        2019: 0.2907884564161656
    }
}

highest_values = {}
#find highest quotient (used in heuristic)
for prod, ye in quotients.items():
    highest_value = max(ye.values())
    highest_values[prod] = highest_value

#implementation of class Node
class Node:
    def __init__(self, wilaya: Wilaya, parent=None, action=None, path_cost=0, heuristic = 0):
        self.wilaya = wilaya
        self.parent = parent
        self.action = action
        self.path_cost = path_cost
        self.heuristic = heuristic

    def __hash__(self):
        return hash(self.wilaya)

    def __eq__(self, other):
        return isinstance(other, Node) and self.wilaya == other.wilaya
    def __lt__(self, other):
        return (self.heuristic + self.path_cost< other.heuristic + other.path_cost)
    def __str__(self):
        return self.wilaya.get_wilaya_name() 

#implementation of the lowest_price problem class
class ProductPriceProblem:
    def __init__(self, initial_wilaya: Wilaya, product: product,year,season):
        self.initial_wilaya = initial_wilaya
        self.year=year
        self.season=season
        self.product = product
        self.lowest_price = float('inf')  # Initialize with a very high value to do comparison
        self.wilayet = wilayas
        self.nei = wilaya_neighbors
        self.G = nx.Graph(self.nei)

    #function that return neighbors of a current wilaya
    def possible_directions(self, wilaya):
        if wilaya in self.G:
            return list(self.G.neighbors(wilaya))
        else:
            return []

    def get_initial_state(self):
        return self.initial_wilaya
    
    #heuristic function for lowest_price  problem
    def heuristic_abzk(self,Wilaya,product,year):
        year=self.year
        product=self.product.get_name()
        production = self.wilayet[Wilaya.get_wilaya_number()].wilaya_products[product].yearly_info.yearly_production_data[year][1]
        mean= mean_production[product][year]
        quotient= mean/production
        return quotient
    
    #function which return whether a node is the goal or not
    def is_goal_state(self, wilaya):
        year=self.year
        product=self.product
        heuristic_quotient = self.heuristic_abzk(wilaya,product,year)
        #target_quotient = quotients[self.product.get_name()][year]
        #print(f"quotient : {heuristic_quotient}")
        if heuristic_quotient <= highest_values[product.get_name()]:
            return True
        else:
            return False
    def is_goal_state_a_star(self, node):
        year=self.year
        product=self.product
        heuristic_quotient = self.heuristic_abzk(node.wilaya,product,year)
        #target_quotient = quotients[self.product.get_name()][year]
        #print(f"quotient : {heuristic_quotient}")
        if heuristic_quotient <= highest_values[product.get_name()]:
            return True
        else:
            return False
     
    #function return cost(distance in KM)  between wilaya and one of it's neighbors
    def get_step_cost(self, start_wilaya , end_wilaya):
        z = 0
        for i in range(0, len(cost[start_wilaya]['Neighbors'])):
        
            if cost[start_wilaya]['Neighbors'][i][0] == end_wilaya:
               z = i 
            break
        return cost[start_wilaya]['Neighbors'][z][1]
    
    #function which return the best neighbor of any wilaya according to their prices level

    def best_neighbors(self,node):
        successors = self.possible_directions(node.wilaya)
        best_neighbor = None
        lowest_price = float('inf')
        for successor in successors:
            price =  successor.wilaya_products[self.product.get_name()].yearly_info.get_prices_per_season(self.year,self.season)
            if price < lowest_price:
                lowest_price = price
                best_neighbor = successor
        b = Node(best_neighbor,node,lowest_price)
        return b

#breath first search algorithm
def bfs(problem):
    start = problem.get_initial_state()
    visited = set()
    node = Node(problem.get_initial_state())
    queue = deque([node])

    while queue:
     node = queue.popleft()
     while node in visited:
        if not len(queue)==0:
            node = queue.popleft()
        else:
            return None
     #print("Visiting node:", node.wilaya.get_wilaya_number())  # Add this line for debugging
     if problem.is_goal_state(node.wilaya):
       return node.wilaya.get_wilaya_name(),node.wilaya.wilaya_products[problem.product.get_name()].yearly_info.get_prices_per_season(problem.year,problem.season)
     visited.add(node)
     for child in problem.possible_directions(node.wilaya):
        child = Node(child, node)
        queue.append(child)
        node =queue.popleft()
        if node not in visited:
            queue.append(node)
    return None

# depth first search algorithm
def dfs(problem):
    start = problem.get_initial_state()
    visited = set()
    node = Node(problem.get_initial_state())
    frontier = [node]
    while frontier:
     node =frontier.pop()
     #print("Visiting node:", node.wilaya.get_wilaya_number())  # Add this line for debugging
     if problem.is_goal_state(node.wilaya):
       return node.wilaya.get_wilaya_name(),node.wilaya.wilaya_products[problem.product.get_name()].yearly_info.get_prices_per_season(problem.year,problem.season)
     visited.add(node)
     for child in problem.possible_directions(node.wilaya):
        child = Node(child, node)
        frontier.append(child)
        node =frontier.pop()
        if node not in visited:
            frontier.append(node)
    return None
# hill climbing / type : steepest acent 
def hill_climbing(problem):
    currentState = Node(problem.get_initial_state())
    while True:
        neighbor_state = problem.best_neighbors(currentState)
        if neighbor_state.wilaya.wilaya_products[problem.product.get_name()].yearly_info.get_prices_per_season(problem.year,problem.season) > currentState.wilaya.wilaya_products[problem.product.get_name()].yearly_info.get_prices_per_season(problem.year,problem.season):
            return currentState.wilaya.get_wilaya_name()
        else :
            currentState = neighbor_state
            #print("yoo ",currentState.wilaya.get_wilaya_name(),currentState.wilaya.wilaya_products[problem.product.get_name()].yearly_info.get_prices_per_season(problem.year,problem.season))

#A* search algorithm 
def AStarSearch(problem):
    node = Node(problem.get_initial_state())
    closed = set()
    opened =  PriorityQueue()
    opened.put(node) # Priority queue with initial cost, start node, and path

    while opened:
        node = opened.get()
        
        if problem.is_goal_state_a_star(node):
            print('goal:'+node.wilaya.get_wilaya_name()) #print the goal node name
            return node.wilaya.get_wilaya_name(),node.wilaya.wilaya_products[problem.product.get_name()].yearly_info.get_prices_per_season(problem.year,problem.season)
        closed.add(node)

        # Expand the current node's neighbors
        for neighbor in problem.nei[node.wilaya]:
                g_cost = problem.get_step_cost(node.wilaya.get_wilaya_name(),neighbor.get_wilaya_name())
                h_cost = problem.heuristic_abzk(neighbor, problem.product.get_name(),problem.year)
                node = Node(neighbor,heuristic=h_cost,path_cost=g_cost,parent=node)
                if not node in closed:
                    opened.put(node)

    return None


w = wilayas[31]
pr = product("aubergine")
s="winter"

p = ProductPriceProblem(w,pr,2016,s)

# Directory paths
csv_directory1 = "C:/python/projet/data_projet/data_2016-2019_csv/"
csv_directory = "C:/python/projet/data_projet/prices_data_2016-2019_csv/"
csv_directory3 = "C:/python/projet/data_projet/sorted_data_csv/"
csv_directory2 = "C:\python\projet\data_projet\consumption.csv"
csv_directory4 = "C:\python\projet\data_projet\production.csv"
'''
# Ensure the consumption directory exists
if not os.path.exists(csv_directory2):
    os.makedirs(csv_directory2)
if not os.path.exists(csv_directory4):
    os.makedirs(csv_directory4)
'''
# Lock file paths
lock_file_path1 = lambda product_name, year: f"{csv_directory1}{product_name}_{year}.csv.lock"
lock_file_path = lambda product_name, year: f"{csv_directory}{product_name}_{year}.csv.lock"
lock_file_path3 = lambda product_name, year: f"{csv_directory3}{product_name}_{year}.csv.lock"
lock_file_path2 = f"{csv_directory2}.csv.lock"
lock_file_path4 = f"{csv_directory4}.csv.lock"

# Read and process production data
for product_name in product_names:
    for year in years:
        number=1
        file_path1 = f"{csv_directory1}{product_name}_{year}.csv"
        lock = FileLock(lock_file_path1(product_name, year))

        with lock, open(file_path1, newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile, delimiter=',')
            #print(f"Reading data for {product_name} - {year}:")
            file_processed = False
            for row in reader:
                try:
                    if len(row) < 4:
                        print(f"Warning: Incomplete data in CSV file: {row}")
                        continue
                    wilaya_name = row[0]
                    land_size = float(row[1])
                    yearly_production = float(row[2])
                    productivity = float(row[3])
                    if land_size == 0 and yearly_production == 0 and productivity ==0:
                        land_size = 1
                        yearly_production = 1
                        productivity = 1
                    else:
                        land_size = float(row[1])
                        yearly_production = float(row[2])
                        productivity = float(row[3])
                    #print(f"Wilaya: {wilaya_name} , land_size: {land_size} , yearly_production: {yearly_production} , productivity: {productivity} ")

                    p.wilayet[number].wilaya_products[product_name].yearly_info.set_yearly_production(year, land_size, yearly_production, productivity)
                    number += 1
                    file_processed = True
                except ValueError as e:
                    print(f"Error converting row to floats: {row}")
                    print(e)

            if not file_processed:
                print(f"No data found for {product_name} - {year}")
            #print("-" * 40)

# Read and process price data
for product_name in product_names:
    for year in years:
        number=1
        file_path = f"{csv_directory}{product_name}_{year}.csv"
        lock = FileLock(lock_file_path(product_name, year))

        with lock, open(file_path, newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile, delimiter=',')
            file_processed = False
            for row in reader:
                try:
                    if len(row) < 5:
                        print(f"Warning: Incomplete data in CSV file: {row}")
                        continue

                    wilaya_name = row[0]
                    season_prices = [float(price) for price in row[1:5]]

                    for i, season in enumerate(seasons):
                        p.wilayet[number].wilaya_products[product_name].yearly_info.set_prices_per_season(year, season_prices[i], season)
                    number += 1
                    file_processed = True
                except ValueError as e:
                    print(f"Error converting row to floats: {row}")
                    print(e)

            if not file_processed:
                print(f"No data found for {product_name} - {year}")
# Directory paths

csv_directory2 = "C:\python\projet\data_projet\consumption.csv"
csv_directory4 = "C:\python\projet\data_projet\production.csv"

lock_file_path2 = f"{csv_directory2}.csv.lock"
lock_file_path4 = f"{csv_directory4}.csv.lock"



# Dictionary to store product instances
products_dict = {name: product(name) for name in product_names}

# Read and process consumption data
lock = FileLock(lock_file_path2)
with lock, open(csv_directory2, newline='', encoding='utf-8') as csvfile:
    reader = csv.reader(csvfile, delimiter=',')
    for row in reader:
        if row:
            product_name = row[0]
            consumption_data = [int(x) for x in row[1:]]
            current_product = products_dict[product_name]
            for i in range(4):
                year = years[i]
                consumption = consumption_data[i]
                current_product.yearly_info.set_country_yearly_consumption(year, consumption)

# Read and process production data
lock = FileLock(lock_file_path4)
with lock, open(csv_directory4, newline='', encoding='utf-8') as csvfile:
    reader = csv.reader(csvfile, delimiter=',')
    for row in reader:
        if row:
            product_name = row[0]
            production_data = [int(x) for x in row[1:]]
            current_product = products_dict[product_name]
            for i in range(4):
                year = years[i]
                production = production_data[i]
                current_product.yearly_info.set_country_yearly_production(year, production)


# Measure time in microseconds with multiple trials
def measure_time_bfs(trials=100):
    times = []
    for _ in range(trials):
        start_time = time.perf_counter()
        bfs(p)
        end_time = time.perf_counter()
        time_taken = (end_time - start_time) * 1e6  # convert to microseconds
        times.append(time_taken)
    avg_time = sum(times) / len(times)
    return avg_time
# Measure space in bytes
tracemalloc.start()
bfs(p)
current, peak = tracemalloc.get_traced_memory()
tracemalloc.stop()

avg_time = measure_time_bfs()
print(f"Average time taken over {100} trials: {avg_time:.2f} µs")
print(f"Current memory usage: {current} bytes; Peak: {peak} bytes")


# Measure time in microseconds with multiple trials
def measure_time_dfs(trials=100):
    times = []
    for _ in range(trials):
        start_time = time.perf_counter()
        dfs(p)
        end_time = time.perf_counter()
        time_taken = (end_time - start_time) * 1e6  # convert to microseconds
        times.append(time_taken)
    avg_time = sum(times) / len(times)
    return avg_time

# Measure space in bytes
tracemalloc.start()
dfs(p)
current, peak = tracemalloc.get_traced_memory()
tracemalloc.stop()

avg_time = measure_time_dfs()
print(f"Average time taken over {100} trials: {avg_time:.2f} µs")
print(f"Current memory usage: {current} bytes; Peak: {peak} bytes")

# Measure time in microseconds with multiple trials
def measure_time_a_star(trials=100):
    times = []
    for _ in range(trials):
        start_time = time.perf_counter()
        AStarSearch(p)
        end_time = time.perf_counter()
        time_taken = (end_time - start_time) * 1e6  # convert to microseconds
        times.append(time_taken)
    avg_time = sum(times) / len(times)
    return avg_time

# Measure space in bytes
tracemalloc.start()
AStarSearch(p)
current, peak = tracemalloc.get_traced_memory()
tracemalloc.stop()

avg_time = measure_time_a_star()
print(f"Average time taken over {100} trials: {avg_time:.2f} µs")
print(f"Current memory usage: {current} bytes; Peak: {peak} bytes")


# Measure time in microseconds with multiple trials
def measure_time_hill_climbing(trials=100):
    times = []
    for _ in range(trials):
        start_time = time.perf_counter()
        hill_climbing(p)
        end_time = time.perf_counter()
        time_taken = (end_time - start_time) * 1e6  # convert to microseconds
        times.append(time_taken)
    avg_time = sum(times) / len(times)
    return avg_time

# Measure space in bytes
tracemalloc.start()
hill_climbing(p)
current, peak = tracemalloc.get_traced_memory()
tracemalloc.stop()

avg_time = measure_time_hill_climbing()
print(f"Average time taken over {100} trials: {avg_time:.2f} µs")
print(f"Current memory usage: {current} bytes; Peak: {peak} bytes")